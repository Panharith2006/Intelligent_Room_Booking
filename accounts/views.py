
from django.db import models
from django.conf import settings
# ============================================================================
# IMPORTS - Consolidated and deduplicated
# ============================================================================
# Django core imports
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth import login, authenticate, logout, get_user_model, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import Group
from django.contrib.auth.hashers import check_password
from django.db.models import Q, Count
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.views.decorators.http import require_GET
from django.views.decorators.csrf import csrf_exempt

# Python standard library
from functools import wraps
import re


# Third-party imports
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

# Local imports
from booking.models import Room, Booking, Announcement
from accounts.forms import UserUpdateForm

# Get the custom User model
User = get_user_model()
# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def get_user_role(user):
    """Return 'Admin' if user is in Admin group, else 'User'."""
    if not user.is_authenticated:
        return 'Unauthenticated'
    
    if user.groups.filter(name='Admin').exists():
        return 'Admin'
    elif user.groups.filter(name='User').exists():
        return 'User'
    else:
        # Default to User if no group assigned
        assign_user_role(user, 'User')
        return 'User'

# ============================================================================
# AJAX ENDPOINT FOR ROOM LIST
# ============================================================================

@require_GET
def ajax_room_list(request):
    """AJAX endpoint for real-time room list"""
    rooms = Room.objects.all().order_by('room_number')
    room_list = []
    for room in rooms:
        room_list.append({
            'id': room.id,
            'name': room.name,
            'room_number': room.room_number,
            'capacity': room.capacity,
            'room_type': room.room_type,
            'is_available': room.is_available,
            'description': room.description,
            'equipment': room.equipment,
            'image_url': room.image.url if hasattr(room, 'image') and room.image else '',
        })
    return JsonResponse({'rooms': room_list})



@login_required
def booked_view(request):
    """View for showing user's bookings: upcoming and history."""
    now = timezone.now()
    upcoming_bookings = Booking.objects.filter(
        user=request.user,
        end_time__gte=now
    ).order_by('start_time')
    past_bookings = Booking.objects.filter(
        user=request.user,
        end_time__lt=now
    ).order_by('-start_time')

    # Active bookings are upcoming confirmed bookings
    active_count = upcoming_bookings.filter(status='confirmed').count()
    all_user_bookings = Booking.objects.filter(user=request.user)
    confirmed_count = all_user_bookings.filter(status='confirmed').count()
    cancelled_count = all_user_bookings.filter(status='cancelled').count()
    total_bookings = all_user_bookings.count()

    context = {
        'user': request.user,
        'bookings': upcoming_bookings,
        'past_bookings': past_bookings,
        'active_count': active_count,
        'confirmed_count': confirmed_count,
        'cancelled_count': cancelled_count,
        'total_bookings': total_bookings,
    }
    return render(request, 'UserPage/booked.html', context)


@login_required
def room_schedule_view(request):
    """Stub view for room schedule page. Replace with real logic as needed."""
    return render(request, 'UserPage/room_schedule.html')

@csrf_exempt
@require_GET
def api_room_bookings(request):
    """
    Return all bookings as JSON for FullCalendar. This endpoint returns all booking history for all users, not just the current user. Supports filtering by room and date range for calendar UI.
    """
    room_id = request.GET.get('room_id')
    start = request.GET.get('start')
    end = request.GET.get('end')

    # Get all bookings (not filtered by user)
    qs = Booking.objects.select_related('room', 'user').all()
    if room_id:
        qs = qs.filter(room_id=room_id)
    if start and end:
        qs = qs.filter(start_time__lt=end, end_time__gt=start)
    events = []
    for booking in qs:
        # Robust building/location extraction
        building = getattr(booking.room, 'room_number', '')
        location = getattr(booking.room, 'location', '')
        if not building and location:
            import re
            match = re.search(r'(Building|Bldg)\s*[A-Za-z0-9]+', location)
            if match:
                building = match.group(0)
            else:
                building = location.split(',')[0].strip()
        events.append({
            'id': booking.id,
            'title': f"{booking.room.name} ({booking.room.room_number}) - {booking.user.get_full_name() or booking.user.username}",
            'start': booking.start_time.isoformat(),
            'end': booking.end_time.isoformat(),
            'room': booking.room.room_number,
            'room_id': booking.room.id,
            'room_name': booking.room.name,
            'room_number': booking.room.room_number,
            'location': location,
            'status': booking.status,
            'user': booking.user.get_full_name() or booking.user.username,
            'user_id': booking.user.id,
            'purpose': booking.purpose,
        })
    return JsonResponse(events, safe=False)



# ============================================================================
# ACCESS CONTROL DECORATORS AND MIDDLEWARE
# ============================================================================

def admin_required(view_func):
    """Decorator to ensure only admins can access admin views"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, 'Please login to access this page.')
            return redirect('accounts:login')
        
        user_role = get_user_role(request.user)
        if user_role != 'Admin':
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('accounts:user_dashboard')
        
        return view_func(request, *args, **kwargs)
    return wrapper

def user_required(view_func):
    """Decorator to ensure only regular users can access user views"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, 'Please login to access this page.')
            return redirect('accounts:login')
        
        user_role = get_user_role(request.user)
        if user_role == 'Admin':
            messages.info(request, 'Redirecting to admin dashboard.')
            return redirect('accounts:admin_dashboard')
        
        return view_func(request, *args, **kwargs)
    return wrapper

def role_redirect(view_func):
    """Decorator to redirect based on user role"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, 'Please login to access this page.')
            return redirect('accounts:login')
        
        user_role = get_user_role(request.user)
        if user_role == 'Admin':
            return redirect('accounts:admin_dashboard')
        else:
            return redirect('accounts:user_dashboard')

    return wrapper

# ============================================================================
# AUTHENTICATION VIEWS
# ============================================================================


def register(request):
    """User registration view - Works with your existing HTML"""
    if request.method == 'POST':
        try:
            # Get form data exactly as your HTML sends it
            first_name = request.POST.get('firstName', '').strip()
            last_name = request.POST.get('lastName', '').strip()
            email = request.POST.get('email', '').strip()
            faculty = request.POST.get('faculty', '').strip()
            department = request.POST.get('department', '').strip()
            password = request.POST.get('password', '')
            confirm_password = request.POST.get('confirmPassword', '')
            student_id = request.POST.get('studentId', '').strip()
            phone_number = request.POST.get('phoneNumber', '').strip()
            
            # Basic validation
            if not all([first_name, last_name, email, password, confirm_password]):
                messages.error(request, 'Please fill in all required fields.')
                return render(request, 'SignIn-RegisterPage/register.html')
            
            if password != confirm_password:
                messages.error(request, 'Passwords do not match.')
                return render(request, 'SignIn-RegisterPage/register.html')
            
            if User.objects.filter(email=email).exists():
                messages.error(request, 'A user with this email already exists.')
                return render(request, 'SignIn-RegisterPage/register.html')
            
            # Create user with flexible fields
            user = User.objects.create_user(
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                student_id=student_id or " ",
                phone_number=phone_number or "000-000-0000",
                faculty=faculty,
                department=department
            )
            
            # Setup user role
            from django.contrib.auth.models import Group
            user_group, created = Group.objects.get_or_create(name='User')
            user.groups.add(user_group)
            
            messages.success(request, 'Registration successful! You can now log in.')
            return redirect('accounts:login')
            
        except Exception as e:
            print(f"Registration error: {e}")
            messages.error(request, f'Registration failed: {str(e)}')
            return render(request, 'SignIn-RegisterPage/register.html')
    
    return render(request, 'SignIn-RegisterPage/register.html')

def custom_login_view(request):
    """Login view - Routes to appropriate dashboard based on account type"""
    if request.method == 'POST':
        email = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        selected_role = request.POST.get('user_role', '').strip()

        print(f"Login attempt for: {email} as {selected_role}")

        # Try authenticating with `username` first (ModelBackend expects this),
        # then fall back to `email` for backends that accept it (e.g. allauth).
        user = authenticate(request, username=email, password=password)
        if user is None:
            user = authenticate(request, email=email, password=password)
        if user is not None:
            # Check group membership
            from django.contrib.auth.models import Group
            is_admin = user.groups.filter(name='Admin').exists()
            is_user = user.groups.filter(name='User').exists()

            # Validate role selection
            if selected_role == 'admin':
                if is_admin:
                    login(request, user)
                    messages.success(request, f'Welcome Admin, {user.first_name}!')
                    print("Admin login - redirecting to admin dashboard")
                    return redirect('accounts:admin_dashboard')
                else:
                    messages.error(request, 'You do not have admin privileges. Please select the correct account type.')
            elif selected_role == 'guest':
                if is_user:
                    login(request, user)
                    messages.success(request, f'Welcome back, {user.first_name}!')
                    print("User login - redirecting to user dashboard")
                    return redirect('accounts:user_dashboard')
                else:
                    messages.error(request, 'You do not have a user account. Please select the correct account type.')
            else:
                messages.error(request, 'Please select your account type.')
        else:
            print("Authentication failed")
            messages.error(request, 'Invalid email or password.')

    return render(request, 'SignIn-RegisterPage/login.html')

def custom_logout_view(request):
    """Logout view"""
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('accounts:login')

# ============================================================================
# USER VIEWS (All UserPage Templates) - FOR STUDENTS
# ============================================================================

@login_required
@user_required
def user_dashboard_view(request):
    """User dashboard - UserPage/featureRoom.html"""
    user_role = get_user_role(request.user)
    
    # Show all rooms (available and unavailable) in featureRoom.html
    from booking.models import Room, Announcement
    from django.utils import timezone
    
    rooms = Room.objects.all().order_by('room_number')
    # Convert rooms to format expected by frontend
    rooms_data = []
    for room in rooms:
        rooms_data.append({
            'id': room.id,
            'name': room.name,
            'room_number': room.room_number,
            'capacity': room.capacity,
            'room_type': room.room_type,
            'available': room.is_available and getattr(room, 'availability_status', 'available') == 'available',
            'is_available': room.is_available,
            'availability_status': getattr(room, 'availability_status', 'available'),
            'description': room.description or f"Modern {room.get_room_type_display().lower()} with capacity for {room.capacity} people.",
            'image_url': room.image.url if getattr(room, 'image', None) else '',
            'equipment': room.equipment or '',
        })
    
    # Get active announcements for users
    announcements = Announcement.objects.filter(
        is_active=True
    ).filter(
        Q(show_until__isnull=True) | Q(show_until__gte=timezone.now())
    ).order_by('-priority', '-created_at')[:5]  # Show top 5 announcements
    
    context = {
        'user': request.user,
        'user_role': user_role,
        'full_name': f"{request.user.first_name} {request.user.last_name}",
        'rooms_data': rooms_data,  
        'announcements': announcements,  
    }
    
    return render(request, 'UserPage/featureRoom.html', context)

@login_required
@user_required
def booking_view(request):
    """Room booking - UserPage/booking.html - Integration with booking app"""
    user_role = get_user_role(request.user)
    
    # Get URL parameters for autofill
    room_id = request.GET.get('room_id')
    date_param = request.GET.get('date')
    time_param = request.GET.get('time')
    
    # Import booking models
    from booking.models import Room, Booking
    
    # Get available rooms for the form (for booking selection)
    rooms = Room.objects.filter(is_available=True).order_by('room_number')

    # Build room_building_map: {room_id: building_code}
    room_building_map = {str(room.id): room.room_number for room in rooms}

    # Get selected room details if room_id is provided
    selected_room = None
    if room_id:
        try:
            selected_room = Room.objects.get(id=room_id, is_available=True)
        except Room.DoesNotExist:
            selected_room = None

    # Get today's date for form minimum date
    today = timezone.now().date()

    # Get user's booking statistics
    user_bookings = Booking.objects.filter(user=request.user)
    total_bookings = user_bookings.count()

    # All bookings are confirmed or cancelled
    confirmed_bookings = user_bookings.filter(status='confirmed').count()

    # Active bookings are upcoming confirmed bookings
    now = timezone.now()
    active_bookings = user_bookings.filter(status='confirmed', start_time__gte=now).count()

    # Get recent bookings for reference
    recent_bookings = user_bookings.order_by('-created_at')[:5]

    # Get room types for filtering
    room_types = Room.ROOM_TYPES

    # Check if user has reached daily booking limit
    daily_bookings = user_bookings.filter(
        start_time__date=today,
        status__in=['confirmed']
    ).count()

    # can_book_today = daily_bookings < 3  

    # Get active announcements for users
    from booking.models import Announcement
    announcements = Announcement.objects.filter(
        is_active=True
    ).filter(
        Q(show_until__isnull=True) | Q(show_until__gte=timezone.now())
    ).order_by('-priority', '-created_at')[:3]  # Show top 3 announcements on booking page

    context = {
        'user': request.user,
        'user_role': user_role,
        'rooms': rooms,
        'room_types': room_types,
        'total_bookings': total_bookings,
        'active_bookings': active_bookings,
        'confirmed_bookings': confirmed_bookings,
        'recent_bookings': recent_bookings,
        # 'can_book_today': can_book_today,
        'daily_bookings': daily_bookings,
        # 'max_daily_bookings': 3,
        'announcements': announcements, 

        # Autofill parameters
        'selected_room': selected_room,
        'autofill_room_id': room_id,
        'autofill_date': date_param,
        'autofill_time': time_param,
        'room_building_map': room_building_map,
    }

    return render(request, 'UserPage/booking.html', context)
    
# ============================================================================
# USER VIEWS (All UserPage Templates) - FOR ADMINS
# ============================================================================

@login_required
@admin_required
def admin_view_rooms_view(request):
    """View all rooms - AdminPage/viewRooms.html (For admins only)"""
    user_role = get_user_role(request.user)
    
    # Get all rooms for admin
    from booking.models import Room
    rooms = Room.objects.all().order_by('room_number')

    # Add image field for consistency with user view
    rooms_data = []
    for room in rooms:
        rooms_data.append({
            'id': room.id,
            'name': room.name,
            'room_number': room.room_number,
            'capacity': room.capacity,
            'room_type': room.room_type,
            'is_available': room.is_available,
            'description': room.description,
            'equipment': room.equipment,
            'image_url': room.image.url if getattr(room, 'image', None) else '/static/images/default-room.png',
        })

    # Get room statistics
    total_rooms = rooms.count()
    available_rooms = rooms.filter(is_available=True).count()
    unavailable_rooms = rooms.filter(is_available=False).count()

    # Add search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        rooms = rooms.filter(
            Q(name__icontains=search_query) |
            Q(room_number__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    # Filter by room type
    room_type = request.GET.get('room_type', '')
    if room_type:
        rooms = rooms.filter(room_type=room_type)

    # Filter by availability
    availability = request.GET.get('availability', '')
    if availability:
        rooms = rooms.filter(is_available=(availability == 'true'))

    # Get room types for filter dropdown
    room_types = Room.ROOM_TYPES

    # Group rooms by type
    room_types_dict = {}
    for room in rooms:
        room_type = room.room_type
        if room_type not in room_types_dict:
            room_types_dict[room_type] = []
        room_types_dict[room_type].append(room)
    
    context = {
        'user': request.user,
        'user_role': user_role,
        'rooms': rooms_data,
        'total_rooms': total_rooms,
        'available_rooms': available_rooms,
        'unavailable_rooms': unavailable_rooms,
        'room_types': room_types_dict,
        'room_types_choices': room_types,
        'search_query': request.GET.get('search', ''),
        'selected_room_type': request.GET.get('room_type', ''),
        'selected_availability': request.GET.get('availability', ''),
    }
    return render(request, 'AdminPage/viewRooms.html', context)

@login_required
def create_booking(request):
    """Handle booking creation with comprehensive validation"""
    user_role = get_user_role(request.user)
    
    if user_role == 'Admin':
        return redirect('accounts:admin_dashboard')
    
    if request.method == 'POST':
        try:
            from booking.models import Room, Booking
            from datetime import datetime
            from django.utils import timezone
            
            # Get form data
            room_id = request.POST.get('room')
            date_str = request.POST.get('date')
            start_time_str = request.POST.get('start_time')
            end_time_str = request.POST.get('end_time')
            purpose = request.POST.get('purpose', '').strip()
            attendees = request.POST.get('attendees', 1)
            notes = request.POST.get('notes', '').strip()
            
            # Basic validation
            if not all([room_id, date_str, start_time_str, end_time_str, purpose]):
                messages.error(request, 'Please fill in all required fields.')
                return redirect('accounts:booking')
            
            # Parse and validate date and time
            try:
                booking_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                start_time = datetime.strptime(start_time_str, '%H:%M').time()
                end_time = datetime.strptime(end_time_str, '%H:%M').time()
            except ValueError:
                messages.error(request, 'Invalid date or time format.')
                return redirect('accounts:booking')
            
            # Create timezone-aware datetime objects
            start_datetime = timezone.make_aware(datetime.combine(booking_date, start_time))
            end_datetime = timezone.make_aware(datetime.combine(booking_date, end_time))
            
            # Validate booking is at least 1 hour in advance
            now = timezone.now()
            if start_datetime <= now + timezone.timedelta(hours=1):
                messages.error(request, 'Bookings must be made at least 1 hour before the scheduled start time.')
                return redirect('accounts:booking')

            # Validate booking is not more than 14 days in advance
            if start_datetime > now + timezone.timedelta(days=14):
                messages.error(request, 'Rooms can only be booked up to 14 days in advance.')
                return redirect('accounts:booking')

            # Validate booking time is between 7:00am and 8:30pm
            allowed_start = start_datetime.replace(hour=7, minute=0, second=0, microsecond=0)
            allowed_end = start_datetime.replace(hour=20, minute=30, second=0, microsecond=0)
            if not (allowed_start <= start_datetime <= allowed_end and allowed_start <= end_datetime <= allowed_end):
                messages.error(request, 'Bookings are only allowed between 7:00am and 8:30pm for each room.')
                return redirect('accounts:booking')

            # Validate end time is after start time
            if start_datetime >= end_datetime:
                messages.error(request, 'End time must be after start time.')
                return redirect('accounts:booking')

            # Validate booking duration (max 8 hours)
            duration = end_datetime - start_datetime
            if duration.total_seconds() > 8 * 3600:
                messages.error(request, 'Maximum booking duration is 8 hours.')
                return redirect('accounts:booking')

            # Validate minimum booking duration (30 minutes)
            if duration.total_seconds() < 30 * 60:
                messages.error(request, 'Minimum booking duration is 30 minutes.')
                return redirect('accounts:booking')
            
            # Get room and validate
            try:
                room = Room.objects.get(id=room_id)
            except Room.DoesNotExist:
                messages.error(request, 'Selected room does not exist.')
                return redirect('accounts:booking')
            
            if not room.is_available:
                messages.error(request, 'This room is not available for booking.')
                return redirect('accounts:booking')
            
            # Validate attendees count
            try:
                attendees_count = int(attendees)
                if attendees_count <= 0:
                    messages.error(request, 'Number of attendees must be at least 1.')
                    return redirect('accounts:booking')
                if attendees_count > room.capacity:
                    messages.error(request, f'Number of attendees ({attendees_count}) exceeds room capacity ({room.capacity}).')
                    return redirect('accounts:booking')
            except (ValueError, TypeError):
                messages.error(request, 'Invalid number of attendees.')
                return redirect('accounts:booking')
            
            # Check for conflicts
            conflicts = Booking.objects.filter(
                    room=room,
                    start_time__lt=end_datetime,
                    end_time__gt=start_datetime,
                    status__in=['confirmed']
                )
            
            if conflicts.exists():
                conflict_booking = conflicts.first()
                conflict_time = conflict_booking.start_time.strftime('%Y-%m-%d %H:%M')
                messages.error(request, f'This time slot conflicts with an existing booking at {conflict_time}.')
                return redirect('accounts:booking')
            
            # Check daily booking limit (optional)
            daily_bookings = Booking.objects.filter(
                user=request.user,
                start_time__date=booking_date,
                status__in=['confirmed', 'pending']
            ).count()
            
            if daily_bookings >= 3:  # Maximum 3 bookings per day
                messages.error(request, 'You have reached the maximum number of bookings per day (3).')
                return redirect('accounts:booking')
            
            # Create booking
            booking = Booking.objects.create(
                user=request.user,
                room=room,
                start_time=start_datetime,
                end_time=end_datetime,
                purpose=purpose,
                attendees=attendees_count,
                additional_notes=notes,
                status='confirmed'  
            )
            
            # Check if this is a redirect from user dashboard
            referrer = request.META.get('HTTP_REFERER', '')
            from_dashboard = ('user-dashboard' in referrer or 
                            'featureRoom' in referrer or 
                            request.GET.get('from_dashboard') == 'true')
            
            # Success message with booking details (only if not from dashboard)
            if not from_dashboard:
                booking_time = start_datetime.strftime('%Y-%m-%d at %H:%M')
                duration_hours = duration.total_seconds() / 3600
                
                messages.success(request, 
                    f'Room booked successfully! '
                    f'Room: {room.name} ({room.room_number}) | '
                    f'Date: {booking_time} | '
                    f'Duration: {duration_hours:.1f} hours | '
                    f'Status: Confirmed'
                )
            
            return redirect('accounts:booked')
            
        except Exception as e:
            messages.error(request, f'Booking failed: {str(e)}')
            return redirect('accounts:booking')
    
    return redirect('accounts:booking')

@login_required
def create_booking_redirect(request):
    """Redirect booking creation to booking app"""
    user_role = get_user_role(request.user)
    
    if user_role == 'Admin':
        return redirect('accounts:admin_dashboard')
    
    # Redirect to booking app's create_booking view
    return redirect('booking:create_booking')


# ============================================================================
# USER SETTINGS AND PROFILE VIEWS
# ============================================================================
@login_required
@user_required
def setting_view(request):
    """User settings - UserPage/setting.html"""
    user_role = get_user_role(request.user)
    
    # Refresh user from database to get latest data
    request.user.refresh_from_db()
    
    context = {
        'user': request.user,
        'user_role': user_role,
    }
    
    return render(request, 'UserPage/setting.html', context)


@login_required
@user_required
def profile_setting_view(request):
    """Edit user profile - UserPage/profileSetting.html - Enhanced for Google OAuth users"""
    user_role = get_user_role(request.user)
    
    # Check if this is a Google user
    is_google_user = hasattr(request.user, 'socialaccount_set') and request.user.socialaccount_set.filter(provider='google').exists()
    
    from accounts.forms import UserUpdateForm
    if request.method == 'POST':
        # AJAX delete profile picture
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.POST.get('action') == 'delete_profile_picture':
            user = request.user
            user.profile_picture.delete(save=False)
            user.profile_picture = None
            user.save()
            return JsonResponse({
                'success': True,
                'message': 'Profile photo deleted.',
                'profile_picture_url': 'https://via.placeholder.com/300x300/4a90e2/ffffff?text=Profile'
            })
        # AJAX profile update
        elif request.headers.get('x-requested-with') == 'XMLHttpRequest':
            form = UserUpdateForm(request.POST, request.FILES, instance=request.user)
            if form.is_valid():
                try:
                    # Log form data before saving
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.info(f"Form data before save: {form.cleaned_data}")
                    
                    user = form.save()
                    
                    # Refresh user instance from database to ensure latest data
                    user.refresh_from_db()
                    
                    # Log user data after saving
                    logger.info(f"User data after save - Faculty: '{user.faculty}', Department: '{user.department}'")
                    
                    pic_url = user.profile_picture.url if user.profile_picture else 'https://via.placeholder.com/300x300/4a90e2/ffffff?text=Profile'
                    
                    # Enhanced success message for Google users
                    if is_google_user:
                        success_msg = 'Profile updated successfully! Your Google account is now fully customized.'
                    else:
                        success_msg = 'Profile updated successfully!'
                        
                    return JsonResponse({
                        'success': True,
                        'message': success_msg,
                        'profile_picture_url': pic_url,
                        'debug_info': {
                            'faculty': user.faculty,
                            'department': user.department
                        }
                    })
                except Exception as e:
                    return JsonResponse({
                        'success': False, 
                        'message': f'Save failed: {str(e)}',
                        'errors': {'__all__': [str(e)]}
                    })
            else:
                # Detailed error reporting
                error_messages = []
                for field, errors in form.errors.items():
                    for error in errors:
                        if field == '__all__':
                            error_messages.append(error)
                        else:
                            error_messages.append(f"{field.replace('_', ' ').title()}: {error}")
                
                return JsonResponse({
                    'success': False, 
                    'message': 'Please correct the following errors: ' + '; '.join(error_messages),
                    'errors': form.errors.as_json()
                })
        # Normal POST (non-AJAX)
        else:
            form = UserUpdateForm(request.POST, request.FILES, instance=request.user)
            if form.is_valid():
                try:
                    form.save()
                    
                    # Enhanced success message for Google users
                    if is_google_user:
                        messages.success(request, 'Profile updated successfully! Your Google account information has been customized.')
                    else:
                        messages.success(request, 'Profile updated successfully!')
                    return redirect('accounts:setting')
                except Exception as e:
                    messages.error(request, f'Failed to save profile: {str(e)}')
            else:
                # Show specific validation errors
                for field, errors in form.errors.items():
                    for error in errors:
                        if field == '__all__':
                            messages.error(request, error)
                        else:
                            field_name = field.replace('_', ' ').title()
                            messages.error(request, f"{field_name}: {error}")
    else:
        form = UserUpdateForm(instance=request.user)
        
        # Add helpful message for new Google users
        if is_google_user and (not request.user.phone_number or not request.user.faculty or request.user.student_id.startswith('GOOGLE')):
            messages.info(request, 'Complete your profile! You can add your student ID, phone number, faculty, and other details to personalize your account.')
    
    context = {
        'user': request.user,
        'user_role': user_role,
        'form': form,
        'is_google_user': is_google_user,
    }
    return render(request, 'UserPage/profileSetting.html', context)

@login_required
@user_required
def about_us_view(request):
    """About us - UserPage/about-us.html"""
    user_role = get_user_role(request.user)
    
    context = {
        'user': request.user,
        'user_role': user_role,
    }
    
    return render(request, 'UserPage/about-us.html', context)

# ============================================================================
# SUPPORT/SERVICE VIEW
# ============================================================================


@login_required
@user_required
def service_view(request):
    """Service/Support - UserPage/service.html"""
    user_role = get_user_role(request.user)
    if request.method == 'POST':
        print('[DEBUG] service_view POST handler called')
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        print(f"[DEBUG] POST data: name={name}, email={email}, subject={subject}, message={message}")

        # Validate required fields
        if not (name and email and subject and message):
            print('[ERROR] Missing required fields in contact form')
            messages.error(request, 'All fields are required.')
        else:
            try:
                from booking.telegram_notifications import send_support_message_to_telegram
                send_support_message_to_telegram(name, email, subject, message)
                print("[DEBUG] Called send_support_message_to_telegram from service_view")
                messages.success(request, 'Your message has been sent to support. We will contact you soon!')
            except Exception as e:
                print(f"[ERROR] Failed to send Telegram alert: {e}")
                messages.error(request, f'Failed to send message: {str(e)}')

    context = {
        'user': request.user,
        'user_role': user_role,
    }
    return render(request, 'UserPage/service.html', context)

# ============================================================================
# ADMIN VIEWS (All AdminPage Templates) - FOR ADMINS ONLY=====
# ============================================================================

@login_required
@admin_required
def admin_dashboard_view(request):
    """Admin dashboard - AdminPage/adminHomePage.html"""
    user_role = get_user_role(request.user)
    
    # Get admin statistics
    total_users = User.objects.count()
    admin_count = User.objects.filter(groups__name='Admin').count()
    user_count = User.objects.filter(groups__name='User').count()
    
    # Get booking statistics
    total_bookings = Booking.objects.count()
    confirmed_bookings = Booking.objects.filter(status='confirmed').count()
    total_rooms = Room.objects.count()
    
    context = {
        'user': request.user,
        'user_role': user_role,
        'total_users': total_users,
        'admin_count': admin_count,
        'user_count': user_count,
        'total_bookings': total_bookings,
        'confirmed_bookings': confirmed_bookings,
        'total_rooms': total_rooms,
    }
    
    return render(request, 'AdminPage/adminHomePage.html', context)

@login_required
@admin_required
def manage_rooms_view(request):
    user_role = get_user_role(request.user)
    
    # Handle room management
    if request.method == 'POST':
        try:
            from booking.models import Room
            
            action = request.POST.get('action')
            
            if action == 'add_room':
                # Add room logic
                room_name = request.POST.get('room_name')
                room_number = request.POST.get('room_number')
                room_type = request.POST.get('room_type')
                capacity = request.POST.get('capacity')
                description = request.POST.get('description', '')
                equipment = request.POST.get('equipment', '')
                room_image = request.FILES.get('room_image')
                
                if room_name and room_number and room_type and capacity:
                    # Room number duplicates are now allowed
                    room_data = {
                        'name': room_name,
                        'room_number': room_number,
                        'room_type': room_type,
                        'capacity': int(capacity),
                        'description': description,
                        'equipment': equipment,
                        'is_available': True
                    }
                    
                    if room_image:
                        room_data['image'] = room_image
                        
                    Room.objects.create(**room_data)
                    messages.success(request, f'Room "{room_name}" added successfully!')
                else:
                    messages.error(request, 'Please fill in all required fields.')
                
            elif action == 'edit_room':
                # Edit room logic
                room_id = request.POST.get('room_id')
                room_name = request.POST.get('room_name')
                room_number = request.POST.get('room_number')
                room_type = request.POST.get('room_type')
                capacity = request.POST.get('capacity')
                description = request.POST.get('description', '')
                equipment = request.POST.get('equipment', '')
                room_image = request.FILES.get('room_image')
                
                if room_id and room_name and room_number and room_type and capacity:
                    try:
                        room = Room.objects.get(id=room_id)
                        
                        # Check if room number already exists (excluding current room)
                        if Room.objects.filter(room_number=room_number).exclude(id=room_id).exists():
                            messages.error(request, f'Room number "{room_number}" already exists.')
                        else:
                            room.name = room_name
                            room.room_number = room_number
                            room.room_type = room_type
                            room.capacity = int(capacity)
                            room.description = description
                            room.equipment = equipment
                            
                            # Update image if provided
                            if room_image:
                                room.image = room_image
                                
                            room.save()
                            messages.success(request, f'Room "{room_name}" updated successfully!')
                    except Room.DoesNotExist:
                        messages.error(request, 'Room not found.')
                    except ValueError:
                        messages.error(request, 'Invalid capacity value. Please enter a number.')
                else:
                    messages.error(request, 'Please fill in all required fields.')
                
            elif action == 'delete_room':
                # Delete room logic
                room_id = request.POST.get('room_id')
                if room_id:
                    try:
                        room = Room.objects.get(id=room_id)
                        room_name = room.name
                        
                        # Check if room has active bookings
                        active_bookings = room.bookings.filter(
                            status__in=['confirmed'],
                            start_time__gte=timezone.now()
                        )
                        
                        if active_bookings.exists():
                            messages.error(request, f'Cannot delete room "{room_name}". It has active bookings.')
                        else:
                            room.delete()
                            messages.success(request, f'Room "{room_name}" deleted successfully!')
                    except Room.DoesNotExist:
                        messages.error(request, 'Room not found.')
                        
            elif action == 'toggle_availability':
                # Toggle room availability
                room_id = request.POST.get('room_id')
                if room_id:
                    try:
                        room = Room.objects.get(id=room_id)
                        room.is_available = not room.is_available
                        room.save()
                        status = 'available' if room.is_available else 'unavailable'
                        messages.success(request, f'Room "{room.name}" is now {status}!')
                    except Room.DoesNotExist:
                        messages.error(request, 'Room not found.')
                        
        except ValueError as e:
            messages.error(request, f'Invalid input: {str(e)}')
        except Exception as e:
            messages.error(request, f'Room management failed: {str(e)}')
    
    # Get rooms data
    from booking.models import Room
    rooms = Room.objects.all().order_by('room_number')
    
    # Get room types for the form
    room_types = Room.ROOM_TYPES
    
    # Get room statistics
    total_rooms = rooms.count()
    available_rooms = rooms.filter(is_available=True).count()
    unavailable_rooms = rooms.filter(is_available=False).count()
    
    # Add search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        rooms = rooms.filter(
            Q(name__icontains=search_query) |
            Q(room_number__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filter by room type
    room_type_filter = request.GET.get('room_type', '')
    if room_type_filter:
        rooms = rooms.filter(room_type=room_type_filter)
    
    # Filter by availability
    availability_filter = request.GET.get('availability', '')
    if availability_filter:
        rooms = rooms.filter(is_available=(availability_filter == 'true'))
    
    context = {
        'user': request.user,
        'user_role': user_role,
        'rooms': rooms,
        'room_types': room_types,
        'total_rooms': total_rooms,
        'available_rooms': available_rooms,
        'unavailable_rooms': unavailable_rooms,
        'search_query': request.GET.get('search', ''),
        'selected_room_type': request.GET.get('room_type', ''),
        'selected_availability': request.GET.get('availability', ''),
    }
    
    return render(request, 'AdminPage/admin_room_management.html', context)


@login_required
@admin_required
def admin_room_management_view(request):
    """Admin room management with modal forms - AdminPage/admin_room_management.html"""
    user_role = get_user_role(request.user)
    from django.contrib import messages
    from django.db.models import Q

    # Handle room management via AJAX/Modal POST requests
    if request.method == 'POST':
        try:
            from booking.models import Room
            
            action = request.POST.get('action')
            
            if action == 'add_room':
                # Add room logic
                room_name = request.POST.get('room_name')
                room_number = request.POST.get('room_number')
                room_type = request.POST.get('room_type')
                capacity = request.POST.get('capacity')
                description = request.POST.get('description', '')
                equipment = request.POST.get('equipment', '')
                room_image = request.FILES.get('room_image')
                
                if room_name and room_number and room_type and capacity:
                    # Check if room name or room number already exists
                    if Room.objects.filter(name=room_name).exists():
                        messages.error(request, f'Room name "{room_name}" already exists.')
                    else:
                        room_data = {
                            'name': room_name,
                            'room_number': room_number,
                            'room_type': room_type,
                            'capacity': int(capacity),
                            'description': description,
                            'equipment': equipment,
                            'is_available': True
                        }
                        
                        if room_image:
                            room_data['image'] = room_image
                            
                        Room.objects.create(**room_data)
                        messages.success(request, f'Room "{room_name}" added successfully!')
                else:
                    messages.error(request, 'Please fill in all required fields.')
                
            elif action == 'edit_room':
                # Edit room logic
                room_id = request.POST.get('room_id')
                room_name = request.POST.get('room_name')
                room_number = request.POST.get('room_number')
                room_type = request.POST.get('room_type')
                capacity = request.POST.get('capacity')
                description = request.POST.get('description', '')
                equipment = request.POST.get('equipment', '')
                room_image = request.FILES.get('room_image')
                
                if room_id and room_name and room_number and room_type and capacity:
                    try:
                        room = Room.objects.get(id=room_id)
                        
                        # Check if room number already exists (excluding current room)
                        if Room.objects.filter(room_number=room_number).exclude(id=room_id).exists():
                            messages.error(request, f'Room number "{room_number}" already exists.')
                        else:
                            room.name = room_name
                            room.room_number = room_number
                            room.room_type = room_type
                            room.capacity = int(capacity)
                            room.description = description
                            room.equipment = equipment
                            
                            # Update image if provided
                            if room_image:
                                room.image = room_image
                                
                            room.save()
                            messages.success(request, f'Room "{room_name}" updated successfully!')
                    except Room.DoesNotExist:
                        messages.error(request, 'Room not found.')
                    except ValueError:
                        messages.error(request, 'Invalid capacity value. Please enter a number.')
                else:
                    messages.error(request, 'Please fill in all required fields.')
                    
        except ValueError as e:
            messages.error(request, f'Invalid input: {str(e)}')
        except Exception as e:
            messages.error(request, f'Room management failed: {str(e)}')
    
    # Get rooms data
    from booking.models import Room
    rooms = Room.objects.all().order_by('room_number')
    
    # Get room types for the form
    room_types = Room.ROOM_TYPES
    
    # Get room statistics
    total_rooms = rooms.count()
    available_rooms = rooms.filter(is_available=True).count()
    unavailable_rooms = rooms.filter(is_available=False).count()
    
    # Add search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        rooms = rooms.filter(
            Q(name__icontains=search_query) |
            Q(room_number__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filter by room type
    room_type_filter = request.GET.get('room_type', '')
    if room_type_filter:
        rooms = rooms.filter(room_type=room_type_filter)
    
    # Filter by availability
    availability_filter = request.GET.get('availability', '')
    if availability_filter:
        rooms = rooms.filter(is_available=(availability_filter == 'true'))
    
    context = {
        'user': request.user,
        'user_role': user_role,
        'rooms': rooms,
        'room_types': room_types,
        'total_rooms': total_rooms,
        'available_rooms': available_rooms,
        'unavailable_rooms': unavailable_rooms,
        'search_query': request.GET.get('search', ''),
        'selected_room_type': request.GET.get('room_type', ''),
        'selected_availability': request.GET.get('availability', ''),
    }
    
    return render(request, 'AdminPage/admin_room_management.html', context)


@login_required
@admin_required
def add_room_view(request):
    """Admin add room - AdminPage/addRoom.html"""
    user_role = get_user_role(request.user)
    from booking.models import Room
    room_types = Room.ROOM_TYPES if hasattr(Room, 'ROOM_TYPES') else [
        ('classroom', 'Classroom'),
        ('lab', 'Laboratory'),
        ('conference', 'Conference Room'),
        ('auditorium', 'Auditorium'),
        ('library', 'Library Room'),
        ('study', 'Study Room'),
        ('other', 'Other'),
    ]
    from django.contrib import messages
    if request.method == 'POST':
        name = request.POST.get('room_name')
        room_number = request.POST.get('room_number')
        room_type = request.POST.get('room_type')
        capacity = request.POST.get('capacity')
        description = request.POST.get('description', '')
        equipment = request.POST.get('equipment', '')
        image = request.FILES.get('image')

        if not (name and room_number and room_type and capacity):
            messages.error(request, 'Please fill in all required fields.')
        else:
            try:
                capacity_int = int(capacity)
                room = Room(
                    name=name,
                    room_number=room_number,
                    room_type=room_type,
                    capacity=capacity_int,
                    description=description,
                    equipment=equipment,
                )
                if image:
                    room.image = image
                room.save()
                messages.success(request, f'Room "{name}" added successfully!')
                return redirect('accounts:manage_rooms')
            except Exception as e:
                messages.error(request, f'Error adding room: {str(e)}')

    context = {
        'user': request.user,
        'user_role': user_role,
        'room_types': room_types,
        'action': 'add'
    }
    return render(request, 'AdminPage/admin_room_form.html', context)

@login_required
@admin_required
def all_bookings_view(request):
    """All bookings - AdminPage/allBookings.html"""
    user_role = get_user_role(request.user)
    
    # Handle booking actions
    if request.method == 'POST':
        try:
            from booking.models import Booking
            
            action = request.POST.get('action')
            booking_id = request.POST.get('booking_id')
            
            if action and booking_id:
                booking = Booking.objects.get(id=booking_id)
                room = booking.room
                if action == 'approve':
                    booking.status = 'confirmed'
                    booking.save()
                    # Mark room as occupied
                    room.availability_status = 'occupied'
                    room.is_available = False
                    room.save()
                    messages.success(request, f'Booking for {booking.room.name} has been approved!')
                elif action == 'reject' or action == 'deny':
                    booking.status = 'cancelled'
                    booking.save()
                    # If no other confirmed bookings for this room, mark as available
                    if not room.bookings.filter(status='confirmed').exclude(id=booking.id).exists():
                        room.availability_status = 'available'
                        room.is_available = True
                        room.save()
                    messages.success(request, f'Booking for {booking.room.name} has been rejected!')
                elif action == 'cancel':
                    booking.status = 'cancelled'
                    booking.save()
                    # If no other confirmed bookings for this room, mark as available
                    if not room.bookings.filter(status='confirmed').exclude(id=booking.id).exists():
                        room.availability_status = 'available'
                        room.is_available = True
                        room.save()
                    messages.success(request, f'Booking for {booking.room.name} has been cancelled!')
                    
        except Booking.DoesNotExist:
            messages.error(request, 'Booking not found.')
        except Exception as e:
            messages.error(request, f'Booking action failed: {str(e)}')
    
    # Get all bookings with filtering
    from booking.models import Booking
    from django.utils import timezone
    from datetime import datetime
    import openpyxl
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    bookings = Booking.objects.all().order_by('-start_time')

    # Filtering by category
    category = request.GET.get('category', 'all')
    date_str = request.GET.get('date', '')
    today = timezone.localdate() if hasattr(timezone, 'localdate') else timezone.now().date()

    if category == 'today':
        bookings = bookings.filter(start_time__date=today)
    elif category == 'month':
        # Filter by the current month using a timezone-aware datetime range
        import calendar
        from datetime import time, datetime
        from django.utils import timezone
        first_day = today.replace(day=1)
        last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1])
        start_dt = timezone.make_aware(datetime.combine(first_day, time.min))
        end_dt = timezone.make_aware(datetime.combine(last_day, time.max))
        print(f"[DEBUG] Filtering for month range: {start_dt} to {end_dt} (timezone-aware)")
        filtered = bookings.filter(start_time__gte=start_dt, start_time__lte=end_dt)
        print(f"[DEBUG] All bookings start_time (with year/month, tzinfo, compare to today):")
        for b in bookings:
            print(f"  - {b.start_time} (year={b.start_time.year}, month={b.start_time.month}, tzinfo={b.start_time.tzinfo}) | matches? year: {b.start_time.year == today.year}, month: {b.start_time.month == today.month}")
        print(f"[DEBUG] Filtered bookings:")
        for b in filtered:
            print(f"  - {b.start_time} (year={b.start_time.year}, month={b.start_time.month}, tzinfo={b.start_time.tzinfo})")
        print(f"[DEBUG] Bookings count before filter: {bookings.count()}")
        print(f"[DEBUG] Bookings count after filter: {filtered.count()}")
        if not list(filtered):
            print("[DEBUG] No bookings matched for this month filter!")
        bookings = filtered
    elif category == 'year':
        bookings = bookings.filter(start_time__year=today.year)
    elif category == 'custom' and date_str:
        # Support both 'YYYY-MM-DD' and 'M/D/YYYY' formats
        from datetime import datetime
        import logging
        custom_date = None
        try:
            custom_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            try:
                custom_date = datetime.strptime(date_str, '%m/%d/%Y').date()
            except ValueError:
                pass
        if custom_date:
            from django.utils import timezone
            import datetime as dt
            # Get the start and end of the custom date in the local timezone
            start_dt = timezone.make_aware(dt.datetime.combine(custom_date, dt.time.min))
            end_dt = timezone.make_aware(dt.datetime.combine(custom_date, dt.time.max))
            filtered = bookings.filter(start_time__gte=start_dt, start_time__lte=end_dt)
            # Debug: print all booking start_times and the custom_date
            print(f"[DEBUG] Filtering for custom_date: {custom_date}")
            print(f"[DEBUG] Range: {start_dt} to {end_dt}")
            print(f"[DEBUG] All bookings start_time:")
            for b in bookings:
                print(f"  - {b.start_time}")
            print(f"[DEBUG] Filtered bookings:")
            for b in filtered:
                print(f"  - {b.start_time}")
            bookings = filtered
   

    # Filtering by status
    status = request.GET.get('status', '')
    if status:
        bookings = bookings.filter(status=status)

    # Handle Excel export
    if request.GET.get('export') == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Bookings'
            # Header row
            headers = [
                'ID', 'User', 'Email', 'Room', 'Room Number', 'Start Time', 'End Time', 'Status', 'Purpose', 'Attendees', 'Created At'
            ]
            ws.append(headers)
            for booking in bookings:
                ws.append([
                    booking.id,
                    getattr(booking.user, 'get_full_name', lambda: None)() or booking.user.email,
                    booking.user.email,
                    booking.room.name,
                    booking.room.room_number,
                    booking.start_time.strftime('%Y-%m-%d %H:%M'),
                    booking.end_time.strftime('%Y-%m-%d %H:%M'),
                    booking.status,
                    getattr(booking, 'purpose', ''),
                    getattr(booking, 'attendees', ''),
                    booking.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(booking, 'created_at') else ''
                ])
            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=bookings_export.xlsx'
            wb.save(response)
            return response

    # Calculate statistics
    total_bookings = bookings.count()
    # Active bookings are upcoming confirmed bookings
    now = timezone.now()
    active_bookings = bookings.filter(status='confirmed', start_time__gte=now).count()
    confirmed_bookings = bookings.filter(status='confirmed').count()
    cancelled_bookings = bookings.filter(status='cancelled').count()

    # Get rooms for filtering
    rooms = []
    if bookings.exists():
        rooms = bookings.values_list('room', flat=True).distinct()
        from booking.models import Room
        rooms = Room.objects.filter(id__in=rooms)

    context = {
        'user': request.user,
        'user_role': user_role,
        'bookings': bookings,
        'rooms': rooms,
        'total_bookings': total_bookings,
        'active_bookings': active_bookings,
        'confirmed_bookings': confirmed_bookings,
        'cancelled_bookings': cancelled_bookings,
        'today': today, 
    }
    return render(request, 'AdminPage/allBookings.html', context)

@login_required
@admin_required
def admin_setting_view(request):
    """Admin settings - AdminPage/setting.html"""
    import json

    user_role = get_user_role(request.user)
    form = None
    
    if request.method == 'POST':
        try:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                data = json.loads(request.body)
                user = request.user
                
                # Update user fields
                if 'first_name' in data:
                    user.first_name = data['first_name']
                if 'last_name' in data:
                    user.last_name = data['last_name']
                if 'faculty' in data and hasattr(user, 'faculty'):
                    user.faculty = data['faculty']
                if 'department' in data and hasattr(user, 'department'):
                    user.department = data['department']
                if 'phone_number' in data and hasattr(user, 'phone_number'):
                    user.phone_number = data['phone_number']
                
                # Save the user
                user.save()
                
                # Return success response with updated data
                response_data = {
                    'success': True,
                    'message': 'Profile updated successfully!',
                    'debug_info': {
                        'first_name': user.first_name,
                        'last_name': user.last_name,
                        'email': user.email,
                        'faculty': user.faculty if hasattr(user, 'faculty') else '',
                        'department': user.department if hasattr(user, 'department') else '',
                        'phone_number': user.phone_number if hasattr(user, 'phone_number') else ''
                    }
                }
                return JsonResponse(response_data)
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid data format'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Update failed: {str(e)}'
            }, status=500)
        else:
            # Handle regular form submission
            if form.is_valid():
                try:
                    form.save()
                    messages.success(request, 'Profile updated successfully!')
                    return redirect('accounts:admin_setting')
                except Exception as e:
                    messages.error(request, f'Profile update failed: {str(e)}')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        if field == '__all__':
                            messages.error(request, error)
                        else:
                            field_name = field.replace('_', ' ').title()
                            messages.error(request, f"{field_name}: {error}")

    context = {
        'user': request.user,
        'user_role': user_role,
    }
    return render(request, 'AdminPage/setting.html', context)

@login_required
@admin_required
def manage_users_view(request):
    """Manage users - AdminPage/manageUsers.html"""
    user_role = get_user_role(request.user)
    
    # Handle user management actions
    if request.method == 'POST':
        try:
            action = request.POST.get('action')
            user_id = request.POST.get('user_id')
            
            if action and user_id:
                target_user = User.objects.get(id=user_id)
                
                # Prevent users from modifying themselves
                if target_user.id == request.user.id:
                    messages.error(request, "You cannot modify your own account.")
                    return redirect('accounts:manage_users')
                
                if action == 'make_admin':
                    target_user.is_admin = True
                    target_user.save()
                    messages.success(request, f'User {target_user.get_full_name()} has been made an admin.')
                        
                elif action == 'make_user':
                    target_user.is_admin = False
                    target_user.save()
                    messages.success(request, f'User {target_user.get_full_name()} has been made a regular user.')
                        
                elif action == 'toggle_active':
                    target_user.is_active = not target_user.is_active
                    target_user.save()
                    status = 'activated' if target_user.is_active else 'deactivated'
                    messages.success(request, f'User {target_user.get_full_name()} has been {status}.')
                    
        except User.DoesNotExist:
            messages.error(request, 'User not found.')
        except Exception as e:
            messages.error(request, f'User management failed: {str(e)}')
        
        return redirect('accounts:manage_users')
    
    # Get all users
    try:
        all_users = User.objects.all().order_by('-date_joined')
        admin_users = all_users.filter(is_admin=True)
        regular_users = all_users.filter(is_admin=False)
        
        # Calculate statistics
        total_users = all_users.count()
        active_users = all_users.filter(is_active=True).count()
        inactive_users = total_users - active_users
        admin_count = admin_users.count()
        user_count = regular_users.count()
        
    except Exception as e:
        messages.error(request, f'Error loading users: {str(e)}')
        all_users = []
        admin_users = []
        regular_users = []
        total_users = 0
        active_users = 0
        inactive_users = 0
        admin_count = 0
        user_count = 0
    
    context = {
        'user': request.user,
        'user_role': user_role,
        'all_users': all_users,
        'admin_users': admin_users,
        'regular_users': regular_users,
        'total_users': total_users,
        'active_users': active_users,
        'inactive_users': inactive_users,
        'admin_count': admin_count,
        'user_count': user_count,
    }
    
    return render(request, 'AdminPage/manageUsers.html', context)

@login_required
@admin_required
def admin_room_detail_view(request, room_id):
    """View room details for admin - AdminPage/roomDetail.html"""
    user_role = get_user_role(request.user)
    
    try:
        room = Room.objects.get(id=room_id)
        
        # Get room bookings
        bookings = Booking.objects.filter(room=room).order_by('-start_time')
        
        # Get today's bookings
        today = timezone.now().date()
        today_bookings = bookings.filter(start_time__date=today)
        
        # Get upcoming bookings
        upcoming_bookings = bookings.filter(
            start_time__gt=timezone.now(),
            status__in=['confirmed']
        )[:10]
        
        # Get booking statistics
        total_bookings = bookings.count()
        confirmed_bookings = bookings.filter(status='confirmed').count()
        cancelled_bookings = bookings.filter(status='cancelled').count()
        
        context = {
            'user': request.user,
            'user_role': user_role,
            'room': room,
            'bookings': bookings[:20],  # Limit to 20 recent bookings
            'today_bookings': today_bookings,
            'total_bookings': total_bookings,
            'confirmed_bookings': confirmed_bookings,
            'cancelled_bookings': cancelled_bookings,
        }
        
        return render(request, 'AdminPage/roomDetail.html', context)
        
    except Room.DoesNotExist:
        messages.error(request, 'Room not found.')
        return redirect('accounts:manage_rooms')
    except Exception as e:
        messages.error(request, f'Error loading room details: {str(e)}')
        return redirect('accounts:manage_rooms')

@login_required
@admin_required
def admin_add_room_view(request):
    """Add new room - AdminPage/addRoom.html"""
    user_role = get_user_role(request.user)
    
    try:
        from booking.models import Room
        
        if request.method == 'POST':
            # Get form data
            room_name = request.POST.get('room_name', '').strip()
            room_number = request.POST.get('room_number', '').strip()
            room_type = request.POST.get('room_type', '')
            capacity = request.POST.get('capacity', '')
            description = request.POST.get('description', '').strip()
            equipment = request.POST.get('equipment', '').strip()
            
            # Validation
            if not all([room_name, room_number, room_type, capacity]):
                messages.error(request, 'Please fill in all required fields.')
                return render(request, 'AdminPage/addRoom.html', {
                    'user': request.user,
                    'user_role': user_role,
                    'room_types': Room.ROOM_TYPES,
                })
            
            try:
                capacity = int(capacity)
                if capacity <= 0:
                    messages.error(request, 'Capacity must be a positive number.')
                    return render(request, 'AdminPage/addRoom.html', {
                        'user': request.user,
                        'user_role': user_role,
                        'room_types': Room.ROOM_TYPES,
                    })
            except ValueError:
                messages.error(request, 'Capacity must be a valid number.')
                return render(request, 'AdminPage/addRoom.html', {
                    'user': request.user,
                    'user_role': user_role,
                    'room_types': Room.ROOM_TYPES,
                })
            
            # Check if room number already exists
            if Room.objects.filter(room_number=room_number).exists():
                messages.error(request, f'Room number "{room_number}" already exists.')
                return render(request, 'AdminPage/addRoom.html', {
                    'user': request.user,
                    'user_role': user_role,
                    'room_types': Room.ROOM_TYPES,
                })
            
            # Create room
            room = Room.objects.create(
                name=room_name,
                room_number=room_number,
                room_type=room_type,
                capacity=capacity,
                description=description,
                equipment=equipment,
                is_available=True
            )
            
            messages.success(request, f'Room "{room_name}" ({room_number}) created successfully!')
            return redirect('accounts:manage_rooms')
        
        # GET request - show form
        context = {
            'user': request.user,
            'user_role': user_role,
            'room_types': Room.ROOM_TYPES,
            'action': 'add'
        }
        
        return render(request, 'AdminPage/admin_room_form.html', context)
        
    except Exception as e:
        messages.error(request, f'Error adding room: {str(e)}')
        return redirect('accounts:manage_rooms')

@login_required
@admin_required
def admin_edit_room_view(request, room_id):
    user_role = get_user_role(request.user)
    
    try:
        from booking.models import Room
        room = Room.objects.get(id=room_id)
        
        if request.method == 'POST':
            # Get form data
            room_name = request.POST.get('room_name', '').strip()
            room_number = request.POST.get('room_number', '').strip()
            room_type = request.POST.get('room_type', '')
            capacity = request.POST.get('capacity', '')
            description = request.POST.get('description', '').strip()
            equipment = request.POST.get('equipment', '').strip()
            room_image = request.FILES.get('room_image')  
            
            # Validation
            if not all([room_name, room_number, room_type, capacity]):
                messages.error(request, 'Please fill in all required fields.')
                return render(request, 'AdminPage/admin_room_form.html', {
                    'user': request.user,
                    'user_role': user_role,
                    'room': room,
                    'room_types': Room.ROOM_TYPES,
                    'action': 'edit'
                })
            
            try:
                capacity = int(capacity)
                if capacity <= 0:
                    messages.error(request, 'Capacity must be a positive number.')
                    return render(request, 'AdminPage/admin_room_form.html', {
                        'user': request.user,
                        'user_role': user_role,
                        'room': room,
                        'room_types': Room.ROOM_TYPES,
                        'action': 'edit'
                    })
            except ValueError:
                messages.error(request, 'Capacity must be a valid number.')
                return render(request, 'AdminPage/admin_room_form.html', {
                    'user': request.user,
                    'user_role': user_role,
                    'room': room,
                    'room_types': Room.ROOM_TYPES,
                    'action': 'edit'
                })
            
            # Update room
            room.name = room_name
            room.room_number = room_number
            room.room_type = room_type
            room.capacity = capacity
            room.description = description
            room.equipment = equipment
            
            # Update image if provided
            if room_image:
                room.image = room_image
                
            room.save()
            
            messages.success(request, f'Room "{room_name}" ({room_number}) updated successfully!')
            return redirect('accounts:manage_rooms')
        
        # GET request - show form
        context = {
            'user': request.user,
            'user_role': user_role,
            'room': room,
            'room_types': Room.ROOM_TYPES,
            'action': 'edit'
        }
        
        return render(request, 'AdminPage/admin_room_form.html', context)
        
    except Room.DoesNotExist:
        messages.error(request, 'Room not found.')
        return redirect('accounts:manage_rooms')
    except Exception as e:
        messages.error(request, f'Error editing room: {str(e)}')
        return redirect('accounts:manage_rooms')

@login_required
@admin_required
def admin_delete_room_view(request, room_id):
    """Delete room - AdminPage/deleteRoom.html"""
    user_role = get_user_role(request.user)
    
    try:
        from booking.models import Room, Booking
        room = Room.objects.get(id=room_id)
        
        if request.method == 'POST':
            # Check if room has active bookings
            active_bookings = Booking.objects.filter(
                room=room,
                status__in=['confirmed'],
                start_time__gte=timezone.now()
            )
            
            if active_bookings.exists():
                messages.error(request, f'Cannot delete room "{room.name}". It has {active_bookings.count()} active booking(s).')
                return redirect('accounts:manage_rooms')
            
            # Delete room
            room_name = room.name
            room_number = room.room_number
            room.delete()
            
            messages.success(request, f'Room "{room_name}" ({room_number}) deleted successfully!')
            return redirect('accounts:manage_rooms')
        
        # GET request - show confirmation
        # Get room bookings for display
        bookings = Booking.objects.filter(room=room).order_by('-start_time')
        active_bookings = bookings.filter(
            status__in=['confirmed'],
            start_time__gte=timezone.now()
        )
        
        context = {
            'user': request.user,
            'user_role': user_role,
            'room': room,
            'bookings': bookings[:10],  # Show recent bookings
            'active_bookings': active_bookings,
            'has_active_bookings': active_bookings.exists(),
        }
        
        return render(request, 'AdminPage/room_confirm_delete.html', context)
        
    except Room.DoesNotExist:
        messages.error(request, 'Room not found.')
        return redirect('accounts:manage_rooms')
    except Exception as e:
        messages.error(request, f'Error deleting room: {str(e)}')
        return redirect('accounts:manage_rooms')

# ============================================================================
# AJAX ENDPOINTS FOR ADMIN FUNCTIONALITY
# ============================================================================


@login_required
@admin_required
def ajax_toggle_user_status(request, user_id):
    """AJAX endpoint to toggle user active status"""
    if request.method == 'POST':
        try:
            target_user = User.objects.get(id=user_id)
            
            # Prevent admin from deactivating themselves
            if target_user == request.user:
                if request.headers.get('Content-Type') == 'application/json':
                    return JsonResponse({
                        'success': False, 
                        'error': 'You cannot deactivate your own account.'
                    })
                else:
                    messages.error(request, 'You cannot deactivate your own account.')
                    return redirect('accounts:manage_users')
            
            # Toggle active status
            target_user.is_active = not target_user.is_active
            target_user.save()
            
            status = 'activated' if target_user.is_active else 'deactivated'
            
            # Return JSON for AJAX requests
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({
                    'success': True,
                    'message': f'User {target_user.email} has been {status}.',
                    'user': {
                        'id': target_user.id,
                        'role': get_user_role(target_user),
                        'is_active': target_user.is_active
                    }
                })
            else:
                # Return redirect for form submissions
                messages.success(request, f'User {target_user.email} has been {status}.')
                return redirect('accounts:manage_users')
                
        except User.DoesNotExist:
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({
                    'success': False, 
                    'error': 'User not found.'
                })
            else:
                messages.error(request, 'User not found.')
                return redirect('accounts:manage_users')
        except Exception as e:
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({
                    'success': False, 
                    'error': f'An error occurred: {str(e)}'
                })
            else:
                messages.error(request, f'An error occurred: {str(e)}')
                return redirect('accounts:manage_users')
    
    return JsonResponse({
        'success': False, 
        'error': 'Invalid request method.'
    })

@login_required
@admin_required
def ajax_delete_room(request, room_id):
    """AJAX endpoint to delete a room"""
    if request.method == 'POST':
        try:
            from booking.models import Room
            room = Room.objects.get(id=room_id)
            room_name = room.name
            
            # Check if room has active bookings
            try:
                from booking.models import Booking
                active_bookings = Booking.objects.filter(
                    room=room,
                    status__in=['confirmed'],
                    start_time__gte=timezone.now()
                )
                
                if active_bookings.exists():
                    return JsonResponse({
                        'success': False, 
                        'error': f'Cannot delete room "{room_name}". It has active bookings.'
                    })
            except:
                pass  # If booking model doesn't exist, skip check
            
            room.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Room "{room_name}" has been deleted successfully.'
            })
            
        except Room.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'error': 'Room not found.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False, 
        'error': 'Invalid request method.'
    })

@login_required
@admin_required
def ajax_toggle_room_availability(request, room_id):
    """AJAX endpoint to toggle room availability"""
    if request.method == 'POST':
        try:
            from booking.models import Room
            room = Room.objects.get(id=room_id)
            
            room.is_available = not room.is_available
            room.save()
            
            status = 'available' if room.is_available else 'unavailable'
            
            return JsonResponse({
                'success': True,
                'message': f'Room "{room.name}" is now {status}.',
                'room': {
                    'id': room.id,
                    'name': room.name,
                    'is_available': room.is_available
                }
            })
            
        except Room.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'error': 'Room not found.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False, 
        'error': 'Invalid request method.'
    })

@login_required
@admin_required
def ajax_bulk_action(request):
    """AJAX endpoint for bulk actions on users/rooms"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            action = data.get('action')
            item_ids = data.get('item_ids', [])
            
            if not action or not item_ids:
                return JsonResponse({
                    'success': False, 
                    'error': 'Missing action or item IDs.'
                })
            
            success_count = 0
            error_count = 0
            
            if action in ['make_admin', 'make_user', 'activate_users', 'deactivate_users']:
                # User bulk actions
                for user_id in item_ids:
                    try:
                        target_user = User.objects.get(id=user_id)
                        
                        # Skip current user
                        if target_user == request.user:
                            continue
                            
                        if action == 'make_admin':
                            if assign_user_role(target_user, 'Admin'):
                                success_count += 1
                            else:
                                error_count += 1
                        elif action == 'make_user':
                            if assign_user_role(target_user, 'User'):
                                success_count += 1
                            else:
                                error_count += 1
                        elif action == 'activate_users':
                            target_user.is_active = True
                            target_user.save()
                            success_count += 1
                        elif action == 'deactivate_users':
                            target_user.is_active = False
                            target_user.save()
                            success_count += 1
                            
                    except User.DoesNotExist:
                        error_count += 1
                    except Exception:
                        error_count += 1
                        
            elif action in ['activate_rooms', 'deactivate_rooms', 'delete_rooms']:
                # Room bulk actions
                from booking.models import Room
                for room_id in item_ids:
                    try:
                        room = Room.objects.get(id=room_id)
                        
                        if action == 'activate_rooms':
                            room.is_available = True
                            room.save()
                            success_count += 1
                        elif action == 'deactivate_rooms':
                            room.is_available = False
                            room.save()
                            success_count += 1
                        elif action == 'delete_rooms':
                            # Check for active bookings
                            try:
                                from booking.models import Booking
                                active_bookings = Booking.objects.filter(
                                    room=room,
                                    status__in=['confirmed'],
                                    start_time__gte=timezone.now()
                                )
                                
                                if not active_bookings.exists():
                                    room.delete()
                                    success_count += 1
                                else:
                                    error_count += 1
                            except:
                                room.delete()
                                success_count += 1
                                
                    except Room.DoesNotExist:
                        error_count += 1
                    except Exception:
                        error_count += 1
            
            return JsonResponse({
                'success': True,
                'message': f'Bulk action completed. {success_count} items processed successfully.',
                'details': {
                    'success_count': success_count,
                    'error_count': error_count
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False, 
                'error': 'Invalid JSON data.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False, 
        'error': 'Invalid request method.'
    })

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def setup_user_groups():
    """Setup User and Admin groups"""
    Group.objects.get_or_create(name='User')
    Group.objects.get_or_create(name='Admin')

def assign_user_role(user, role_name):
    """Assign User or Admin role"""
    try:
        user.groups.clear()
        group, created = Group.objects.get_or_create(name=role_name)
        user.groups.add(group)
        
        if role_name == 'Admin':
            user.is_staff = True
            user.is_superuser = True
        else:
            user.is_staff = False
            user.is_superuser = False
        
        user.save()
        print(f"Successfully assigned {role_name} role to {user.email}")
        return True
    except Exception as e:
        print(f"Error assigning role: {e}")
        return False

# ============================================================================
# ADDITIONAL VIEWS
# ============================================================================

@login_required
@role_redirect
def dashboard_view(request):
    """Dashboard router"""
    pass  # This will be handled by the decorator


# ============================================================================
# ADMIN MANAGEMENT FUNCTIONS (For creating admin accounts)
# ============================================================================

def create_admin_account(email, password, first_name, last_name):
    """Function to create admin account - can only be called from Django shell or management command"""
    try:
        if User.objects.filter(email=email).exists():
            print(f"Admin account {email} already exists")
            return False
        
        admin_user = User.objects.create_user(
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            student_id=f"ADM{User.objects.count() + 1:04d}",
            phone_number="000-000-0000",
            faculty="Administration",
            department="IT Department"
        )
        
        setup_user_groups()
        assign_user_role(admin_user, 'Admin')
        
        print(f"Admin account created successfully: {email}")
        return True
        
    except Exception as e:
        print(f"Error creating admin account: {e}")
        return False


@login_required
def user_profile_view(request):
    return redirect('accounts:profile_setting')


@login_required
def booking_detail_view(request, booking_id):
    """View booking details"""
    try:
        from booking.models import Booking
        from datetime import timedelta
        
        booking = Booking.objects.get(id=booking_id, user=request.user)
        
        # Calculate cancellation availability
        current_time = timezone.now()
        time_until_booking = booking.start_time - current_time
        minimum_notice = timedelta(hours=24)
        
        can_cancel = (
            booking.status in ['confirmed'] and 
            time_until_booking >= minimum_notice
        )
        
        context = {
            'booking': booking,
            'user': request.user,
            'user_role': get_user_role(request.user),
            'can_cancel': can_cancel,
            'current_time': current_time,
            'time_until_booking': time_until_booking,
            'hours_until_booking': time_until_booking.total_seconds() / 3600,
        }
        
        return render(request, 'UserPage/booking-detail.html', context)
        
    except Booking.DoesNotExist:
        messages.error(request, 'Booking not found.')
        return redirect('accounts:booked')
    except Exception as e:
        messages.error(request, f'Error viewing booking details: {str(e)}')
        return redirect('accounts:booked')


@login_required
def check_availability_ajax(request):
    """AJAX endpoint to check room availability"""
    if request.method == 'POST':
        try:
            from booking.models import Room, Booking
            import json
            
            data = json.loads(request.body)
            room_id = data.get('room_id')
            date = data.get('date')
            start_time = data.get('start_time')
            end_time = data.get('end_time')
            
            if not all([room_id, date, start_time, end_time]):
                return JsonResponse({'error': 'Missing required parameters'}, status=400)
            
            # Get room
            try:
                room = Room.objects.get(id=room_id)
            except Room.DoesNotExist:
                return JsonResponse({'available': False, 'message': 'Room not found'})
            
            # Check if room is available
            if not room.is_available:
                return JsonResponse({'available': False, 'message': 'Room is not available for booking'})
            
            # Parse date and time
            from datetime import datetime
            try:
                booking_date = datetime.strptime(date, '%Y-%m-%d').date()
                start_time_obj = datetime.strptime(start_time, '%H:%M').time()
                end_time_obj = datetime.strptime(end_time, '%H:%M').time()
                
                start_datetime = datetime.combine(booking_date, start_time_obj)
                end_datetime = datetime.combine(booking_date, end_time_obj)
            except ValueError:
                return JsonResponse({'available': False, 'message': 'Invalid date or time format'})
            
            # Check if booking is in the future
            if start_datetime <= datetime.now():
                return JsonResponse({'available': False, 'message': 'Booking time must be in the future'})
            
            # Check if end time is after start time
            if start_datetime >= end_datetime:
                return JsonResponse({'available': False, 'message': 'End time must be after start time'})
            
            # Check for conflicts
            conflicts = Booking.objects.filter(
                room=room,
                start_time__lt=end_datetime,
                end_time__gt=start_datetime,
                status__in=['confirmed']
            )
            
            if conflicts.exists():
                conflict = conflicts.first()
                conflict_time = conflict.start_time.strftime('%H:%M')
                return JsonResponse({
                    'available': False, 
                    'message': f'Time slot conflicts with existing booking at {conflict_time}',
                    'conflict': {
                        'start_time': conflict.start_time.strftime('%H:%M'),
                        'end_time': conflict.end_time.strftime('%H:%M'),
                        'user': conflict.user.get_full_name()
                    }
                })
            
            # Check daily booking limit
            daily_bookings = Booking.objects.filter(
                user=request.user,
                start_time__date=booking_date,
                status__in=['confirmed']
            ).count()
            
            if daily_bookings >= 3:
                return JsonResponse({
                    'available': False, 
                    'message': 'You have reached the maximum number of bookings per day (3)'
                })
            
            # All checks passed
            return JsonResponse({
                'available': True, 
                'message': 'Room is available for booking',
                'room_info': {
                    'name': room.name,
                    'capacity': room.capacity,
                    'room_type': room.get_room_type_display(),
                    'equipment': room.equipment
                }
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)

@login_required
def get_rooms_ajax(request):
    """AJAX endpoint to get rooms for a specific building"""
    building_id = request.GET.get('building_id')
    room_type = request.GET.get('room_type', '')
    capacity_min = request.GET.get('capacity_min', '')
    
    # Start with available rooms
    rooms = Room.objects.filter(is_available=True)
    
    # Filter by building if provided
    if building_id:
        building_codes = {
            '1': 'A-',
            '2': 'S-',
            '3': 'L-',
            '4': 'B-',
        }
        if building_id in building_codes:
            rooms = rooms.filter(room_number__startswith=building_codes[building_id])
    
    # Filter by room type if provided
    if room_type:
        rooms = rooms.filter(room_type=room_type)
    
    # Filter by minimum capacity if provided
    if capacity_min:
        try:
            capacity_min = int(capacity_min)
            rooms = rooms.filter(capacity__gte=capacity_min)
        except ValueError:
            pass
    
    rooms_data = []
    for room in rooms:
        try:
            # Get next booking for this room
            next_booking = room.bookings.filter(
                start_time__gt=timezone.now(),
                status__in=['confirmed']
            ).order_by('start_time').first()
            
            rooms_data.append({
                'id': room.id,
                'name': room.name,
                'room_number': room.room_number,
                'capacity': room.capacity,
                'room_type': room.room_type,
                'room_type_display': room.get_room_type_display(),
                'description': room.description or '',
                'equipment': room.equipment or '',
                'location': f"{room.name} ({room.room_number})",
                'available': room.is_available,
                'next_booking': next_booking.start_time.strftime('%Y-%m-%d %H:%M') if next_booking else None
            })
        except Exception as e:
            # Skip this room if there's an error (e.g., missing field)
            continue
    
    return JsonResponse({'rooms': rooms_data})

@login_required
def get_buildings_ajax(request):
    """AJAX endpoint to get all buildings - extracted from room numbers"""
    # Extract unique building codes from room numbers
    rooms = Room.objects.all().values_list('room_number', flat=True)
    buildings_set = set()
    
    for room_number in rooms:
        # Extract building code (e.g., "A" from "A-101")
        match = re.match(r'^([A-Z]+)-', str(room_number))
        if match:
            buildings_set.add(match.group(1))
    
    buildings_data = []
    for idx, building_code in enumerate(sorted(buildings_set), 1):
        buildings_data.append({
            'id': idx,
            'name': f'Building {building_code}',
            'code': building_code,
            'address': '',
            'description': f'Building {building_code}'
        })
    
    return JsonResponse({'buildings': buildings_data})


@login_required
@user_required
def cancel_booking_view(request, booking_id):
    """Cancel a booking - with 24-hour advance notice rule"""
    user_role = get_user_role(request.user)
    
    try:
        from booking.models import Booking
        from datetime import timedelta
        
        booking = Booking.objects.get(id=booking_id, user=request.user)
        
        if request.method == 'POST':
            current_time = timezone.now()
            
            # Check if booking status allows cancellation
            if booking.status not in ['confirmed']:
                error_msg = f'Cannot cancel booking with status: {booking.status}'
                messages.error(request, error_msg)
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': error_msg
                    })
                return redirect('accounts:booked')
            
            # Check 24-hour advance notice rule
            time_until_booking = booking.start_time - current_time
            minimum_notice = timedelta(hours=24)
            
            if time_until_booking < minimum_notice:
                hours_left = time_until_booking.total_seconds() / 3600
                if hours_left <= 0:
                    error_msg = 'Cannot cancel booking that has already started or passed.'
                else:
                    error_msg = f'Cannot cancel booking. Must cancel at least 24 hours in advance. Only {hours_left:.1f} hours remaining.'
                
                messages.error(request, error_msg)
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': error_msg
                    })
                return redirect('accounts:booked')
            
            # All checks passed - cancel the booking
            booking.status = 'cancelled'
            booking.save()
            
            # Try to delete Google Calendar event if it exists
            try:
                from booking.google_calendar import GoogleCalendarIntegration
                calendar_integration = GoogleCalendarIntegration()
                if hasattr(booking, 'calendar_event_id') and booking.calendar_event_id:
                    calendar_integration.delete_event(request.user, booking.calendar_event_id)
            except Exception as e:
                print(f"Error deleting calendar event: {e}")
            
            success_msg = f'Booking for {booking.room.name} on {booking.start_time.strftime("%Y-%m-%d at %H:%M")} has been cancelled.'
            messages.success(request, success_msg)
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Booking for {booking.room.name} has been cancelled.'
                })
            
            return redirect('accounts:booked')
        
        # GET request - show confirmation page
        current_time = timezone.now()
        time_until_booking = booking.start_time - current_time
        minimum_notice = timedelta(hours=24)
        
        can_cancel = (
            booking.status in ['confirmed'] and 
            time_until_booking >= minimum_notice
        )
        
        context = {
            'user': request.user,
            'user_role': user_role,
            'booking': booking,
            'can_cancel': can_cancel,
            'current_time': current_time,
            'time_until_booking': time_until_booking,
            'hours_until_booking': time_until_booking.total_seconds() / 3600,
        }
        
        return render(request, 'UserPage/cancelBooking.html', context)
        
    except Booking.DoesNotExist:
        messages.error(request, 'Booking not found.')
        return redirect('accounts:booked')
    except Exception as e:
        messages.error(request, f'Error processing cancellation: {str(e)}')
        return redirect('accounts:booked')
    except Exception as e:
        messages.error(request, f'Error cancelling booking: {str(e)}')
        return redirect('accounts:booked')

@login_required
@user_required
def get_room_details_ajax(request):
    """AJAX endpoint to get room details for autofill"""
    room_id = request.GET.get('room_id')
    
    if not room_id:
        return JsonResponse({'success': False, 'error': 'Room ID is required'})
    
    try:
        
        room = Room.objects.get(id=room_id, is_available=True)
        
        return JsonResponse({
            'success': True,
            'room': {
                'id': room.id,
                'name': room.name,
                'room_number': room.room_number,
                'building_name': room.room_number,
                'capacity': room.capacity,
                'room_type': room.room_type,
            }
        })
    except Room.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Room not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
@login_required
@admin_required
def admin_booking_detail_view(request, booking_id):
    """Admin view for booking details"""
    user_role = get_user_role(request.user)
    
    booking = Booking.objects.get(id=booking_id)
    
    context = {
        'booking': booking,
        'user': request.user,
        'user_role': user_role,
    }
    
    try:
        return render(request, 'AdminPage/booking-detail.html', context)
    except Booking.DoesNotExist:
        messages.error(request, 'Booking not found.')
        return redirect('accounts:all_bookings')
    except Exception as e:
        messages.error(request, f'Error viewing booking: {str(e)}')
        return redirect('accounts:all_bookings')



@login_required
@admin_required
def deactivate_user_view(request):
    """Deactivate user - only for existing admins"""
    user_role = get_user_role(request.user)
    
    if not request.user.is_admin:
        messages.error(request, 'Admin access required.')
        return redirect('accounts:user_dashboard')
    
    if request.method == 'POST':
        try:
            user_email = request.POST.get('user_email')
            action = request.POST.get('action')  # 'deactivate' or 'activate'
            
            if user_email:
                user = User.objects.get(email=user_email)
                
                # Prevent deactivating self
                if user.id == request.user.id:
                    messages.warning(request, 'You cannot deactivate your own account.')
                elif action == 'deactivate':
                    if not user.is_active:
                        messages.warning(request, f'User {user.get_full_name()} is already deactivated.')
                    else:
                        user.is_active = False
                        user.save()
                        messages.success(request, f'User {user.get_full_name()} has been deactivated.')
                elif action == 'activate':
                    if user.is_active:
                        messages.warning(request, f'User {user.get_full_name()} is already active.')
                    else:
                        user.is_active = True
                        user.save()
                        messages.success(request, f'User {user.get_full_name()} has been activated.')
            else:
                messages.error(request, 'Please provide a valid email address.')
                
        except User.DoesNotExist:
            messages.error(request, 'User not found.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        
        return redirect('accounts:deactivate_user')
    
    # Get all users for selection
    active_users = User.objects.filter(is_active=True).exclude(id=request.user.id).order_by('first_name', 'email')
    inactive_users = User.objects.filter(is_active=False).order_by('first_name', 'email')
    
    # Pre-select user if email is provided in GET parameters
    selected_user_email = request.GET.get('user_email', '')
    
    context = {
        'user': request.user,
        'user_role': user_role,
        'active_users': active_users,
        'inactive_users': inactive_users,
        'selected_user_email': selected_user_email,
    }
    
    return render(request, 'AdminPage/deactivateUser.html', context)
